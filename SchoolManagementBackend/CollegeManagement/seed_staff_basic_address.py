import os
import re
from datetime import date, datetime

import openpyxl

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "Swostitech_Acadix.settings")

import django

django.setup()

from django.contrib.auth.hashers import make_password
from django.db import transaction

from Acadix.models import (
    Address,
    Batch,
    Blood,
    Branch,
    City,
    Country,
    EmployeeMaster,
    EmployeeType,
    Gender,
    MotherTongue,
    Nationality,
    Organization,
    Religion,
    State,
    UserLogin,
    UserType,
)
from STAFF.views import get_or_create_default_designation


WORKBOOK_PATH = r"C:\BEC_Production\STAFF (3).xlsx"
ORG_ID = 1
BRANCH_ID = 1
BATCH_ID = 1
CREATED_BY = 1
DEFAULT_PHONE = "1234567890"
DEFAULT_TITLE = "Mr"


def clean_text(value):
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.upper() == "NULL":
        return None
    return re.sub(r"\s+", " ", text)


def normalize_email(value):
    text = clean_text(value)
    if not text:
        return None
    text = text.replace(" @", "@").replace("@ ", "@")
    text = re.sub(r"\s+", "", text)
    return text.lower()


def normalize_phone(value):
    text = clean_text(value)
    if not text:
        return DEFAULT_PHONE
    digits = re.sub(r"\D", "", text)
    return digits if len(digits) == 10 else DEFAULT_PHONE


def normalize_optional_phone(value):
    text = clean_text(value)
    if not text:
        return None
    digits = re.sub(r"\D", "", text)
    return digits if len(digits) == 10 else None


def parse_dob(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = clean_text(value)
    if not text:
        raise ValueError("Date of birth is missing")
    return datetime.strptime(text, "%d.%m.%Y").date()


def parse_title_and_first_name(raw_first_name):
    text = clean_text(raw_first_name) or ""
    title_match = re.match(r"^(Mr|Mrs|Ms)\.?\s*(.*)$", text, flags=re.IGNORECASE)
    if title_match:
        title = title_match.group(1).title()
        first_name = clean_text(title_match.group(2)) or ""
    else:
        title = DEFAULT_TITLE
        first_name = text
    return title, first_name


def parse_address_blob(address_blob):
    text = clean_text(address_blob) or ""
    compact = re.sub(r"\s+", " ", text)
    pattern = re.compile(
        r"Address-(?P<address>.*?),\s*Country-\s*(?P<country>.*?),\s*State-\s*(?P<state>.*?),\s*City/District-\s*(?P<city>.*?),\s*Pincode-\s*(?P<pincode>\d{6})",
        flags=re.IGNORECASE,
    )
    match = pattern.search(compact)
    if not match:
        raise ValueError(f"Unable to parse address: {text}")
    parts = {k: clean_text(v) for k, v in match.groupdict().items()}
    return parts


def map_employee_type(raw_value):
    value = clean_text(raw_value)
    if not value:
        lookup = "Teaching Staff"
    elif "non-teaching" in value.lower():
        lookup = "Non-Teaching Staff"
    else:
        lookup = value
    return EmployeeType.objects.get(employee_type_description__iexact=lookup, is_active=True)


def map_gender(raw_value):
    value = clean_text(raw_value) or ""
    return Gender.objects.get(gender_name__iexact=value, is_active=True)


def map_blood(raw_value):
    value = clean_text(raw_value) or ""
    return Blood.objects.get(blood_name__iexact=value, is_active=True)


def map_nationality(raw_value):
    value = clean_text(raw_value) or ""
    return Nationality.objects.get(nationality_name__iexact=value, is_active=True)


def map_religion(raw_value):
    value = clean_text(raw_value) or ""
    return Religion.objects.get(religion_name__iexact=value, is_active=True)


def map_mother_tongue(raw_value):
    value = clean_text(raw_value) or ""
    return MotherTongue.objects.get(mother_tongue_name__iexact=value, is_active=True)


def map_country(country_name):
    return Country.objects.get(country_name__iexact=country_name, is_active=True)


def map_state(state_name):
    return State.objects.get(state_name__iexact=state_name, is_active=True)


def map_city(city_name):
    return City.objects.get(city_name__iexact=city_name, is_active=True)


def iter_rows():
    wb = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    for r in range(2, ws.max_row + 1):
        values = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if not any(v not in (None, "") for v in values):
            continue
        row = {str(headers[i]): values[i] for i in range(len(headers)) if headers[i] is not None}
        row["_row"] = r
        yield row


def ensure_login(employee, plain_password):
    user_type = UserType.objects.get(id=3, is_active=True)
    login, created = UserLogin.objects.get_or_create(
        user_name=employee.email,
        defaults={
            "plain_password": plain_password,
            "user_type": user_type,
            "reference_id": employee.id,
            "organization": employee.organization,
            "branch": employee.branch,
            "is_active": True,
            "is_staff": True,
            "password": make_password(plain_password),
        },
    )
    if not created:
        login.plain_password = plain_password
        login.user_type = user_type
        login.reference_id = employee.id
        login.organization = employee.organization
        login.branch = employee.branch
        login.is_active = True
        login.is_staff = True
        login.password = make_password(plain_password)
        login.save()
    return login, created


def ensure_address(employee, employee_type, address_parts, phone_value):
    country = map_country(address_parts["country"])
    state = map_state(address_parts["state"])
    city = map_city(address_parts["city"])

    defaults = {
        "organization": employee.organization,
        "branch": employee.branch,
        "usertype": employee_type.employee_type_code,
        "present_address": address_parts["address"],
        "present_pincode": address_parts["pincode"],
        "present_city": str(city.id),
        "present_state": str(state.id),
        "present_country": str(country.id),
        "present_phone_number": phone_value,
        "permanent_address": address_parts["address"],
        "permanent_pincode": address_parts["pincode"],
        "permanent_city": str(city.id),
        "permanent_state": str(state.id),
        "permanent_country": str(country.id),
        "permanent_phone_number": phone_value,
        "created_by": CREATED_BY,
        "updated_by": CREATED_BY,
        "is_active": True,
    }
    address, created = Address.objects.update_or_create(
        reference_id=employee.id,
        defaults=defaults,
    )
    return address, created


def main():
    organization = Organization.objects.get(id=ORG_ID)
    branch = Branch.objects.get(id=BRANCH_ID)
    batch = Batch.objects.get(id=BATCH_ID)
    designation = get_or_create_default_designation(organization, branch, CREATED_BY)

    created_staff = 0
    updated_staff = 0
    created_login = 0
    updated_login = 0
    created_address = 0
    updated_address = 0

    for row in iter_rows():
        with transaction.atomic():
            employee_code = clean_text(row.get("Employee Code"))
            title, first_name = parse_title_and_first_name(row.get("First Name"))
            middle_name = clean_text(row.get("Middle Name"))
            last_name = clean_text(row.get("Last Name"))
            if not title:
                title = DEFAULT_TITLE
            if not first_name:
                raise ValueError(f"Missing first name for row {row['_row']}")

            dob = parse_dob(row.get("Date of Birth"))
            employee_type = map_employee_type(row.get("Employee Type"))
            gender = map_gender(row.get("Gender"))
            blood = map_blood(row.get("Blood Group"))
            nationality = map_nationality(row.get("Nationality"))
            religion = map_religion(row.get("Religion"))
            mother_tongue = map_mother_tongue(row.get("Mother Toungue"))
            email = normalize_email(row.get("Email"))
            if not email:
                raise ValueError(f"Missing email for row {row['_row']}")

            phone_number = normalize_phone(row.get("Mobile No"))
            emergency_contact = normalize_optional_phone(row.get("Emergency Contact No"))
            marital_status = clean_text(row.get("Marital Status")) or "Single"
            address_parts = parse_address_blob(row.get("Address"))

            defaults = {
                "organization": organization,
                "branch": branch,
                "batch": batch,
                "designation": designation,
                "nuid": None,
                "title": title,
                "first_name": first_name,
                "middle_name": middle_name,
                "last_name": last_name,
                "date_of_birth": dob,
                "marital_status": marital_status,
                "gender": gender,
                "nationality": nationality,
                "religion": religion,
                "email": email,
                "phone_number": phone_number,
                "office_email": email,
                "employee_type": employee_type,
                "date_of_joining": date.today(),
                "place_of_birth": address_parts["city"],
                "blood_group": blood,
                "emergency_contact_number": emergency_contact,
                "mother_tongue": mother_tongue,
                "status": "ACTIVE",
                "is_active": True,
                "created_by": CREATED_BY,
            }

            employee, created = EmployeeMaster.objects.update_or_create(
                employee_code=employee_code,
                defaults=defaults,
            )
            if created:
                created_staff += 1
            else:
                updated_staff += 1

            login, login_created = ensure_login(employee, employee.first_name)
            if login_created:
                created_login += 1
            else:
                updated_login += 1

            address, address_created = ensure_address(
                employee,
                employee_type,
                address_parts,
                phone_number,
            )
            if address_created:
                created_address += 1
            else:
                updated_address += 1

            print(
                f"ROW {row['_row']}: employee_code={employee_code} employee_id={employee.id} "
                f"login_id={login.id} address_id={address.id}"
            )

    print(
        f"SUMMARY staff_created={created_staff} staff_updated={updated_staff} "
        f"login_created={created_login} login_updated={updated_login} "
        f"address_created={created_address} address_updated={updated_address}"
    )


if __name__ == "__main__":
    main()
