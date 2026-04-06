from __future__ import annotations

from collections import OrderedDict
from dataclasses import dataclass
from datetime import date
import re
from typing import Any

import openpyxl
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from Acadix.models import AcademicYear, Organization
from Library.models import (
    BookCategory,
    BookLocation,
    BookSubCategory,
    LibraryBook,
    LibraryBooksBarcode,
    LibraryBranch,
    LibraryPurchase,
)


HEADER_ROW = 1


def collapse_spaces(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip()


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return collapse_spaces(value)
    return collapse_spaces(str(value))


def normalize_key(value: Any) -> str:
    return clean_text(value).upper()


def parse_int_or_none(value: Any) -> int | None:
    text = clean_text(value)
    if not text:
        return None
    if re.fullmatch(r"\d+", text):
        return int(text)
    if isinstance(value, (int, float)) and int(value) == value:
        return int(value)
    return None


@dataclass
class GroupedBook:
    title: str
    publisher: str
    author: str
    publish_year: str
    volume: int | None
    isbn: str
    edition: str
    pages: int | None
    copy_count: int = 0


class Command(BaseCommand):
    help = "Import library books from a legacy Excel workbook using grouped book masters."

    def add_arguments(self, parser):
        parser.add_argument("excel_path", type=str)
        parser.add_argument("--created-by", type=int, default=1)
        parser.add_argument("--org-id", type=int, default=1)
        parser.add_argument("--academic-year-id", type=int, default=None)
        parser.add_argument("--dry-run", action="store_true")

    def handle(self, *args, **options):
        excel_path = options["excel_path"]
        created_by = options["created_by"]
        organization_id = options["org_id"]
        academic_year_id = options["academic_year_id"]
        dry_run = options["dry_run"]

        try:
            organization = Organization.objects.get(id=organization_id, is_active=True)
        except Organization.DoesNotExist as exc:
            raise CommandError(f"Organization {organization_id} not found.") from exc

        academic_year = self.resolve_academic_year(organization_id, academic_year_id)
        batch = academic_year.batch

        if LibraryBook.objects.exists() and not dry_run:
            raise CommandError(
                "LibraryBook already contains data. Refusing to import into a non-empty library."
            )

        defaults = self.ensure_defaults(
            organization=organization,
            batch=batch,
            created_by=created_by,
        )

        grouped_books, anomaly_summary = self.read_grouped_books(excel_path)

        self.stdout.write(
            self.style.NOTICE(
                f"Grouped {sum(book.copy_count for book in grouped_books.values())} accession rows "
                f"into {len(grouped_books)} book masters."
            )
        )
        self.stdout.write(self.style.NOTICE(f"Grouping anomalies: {anomaly_summary}"))

        if dry_run:
            self.stdout.write(self.style.SUCCESS("Dry run complete. No data inserted."))
            return

        result = self.import_books(
            grouped_books=grouped_books,
            academic_year=academic_year,
            defaults=defaults,
            created_by=created_by,
        )

        self.stdout.write(self.style.SUCCESS(f"Import complete: {result}"))

    def resolve_academic_year(self, organization_id: int, academic_year_id: int | None) -> AcademicYear:
        queryset = AcademicYear.objects.filter(
            organization_id=organization_id,
            branch_id=1,
            is_active=True,
        ).order_by("id")
        if academic_year_id is not None:
            academic_year = queryset.filter(id=academic_year_id).first()
            if not academic_year:
                raise CommandError(f"Academic year {academic_year_id} not found for organization {organization_id}.")
            return academic_year

        academic_year = queryset.first()
        if not academic_year:
            raise CommandError("No active academic year found.")
        return academic_year

    def ensure_defaults(self, organization, batch, created_by: int) -> dict[str, Any]:
        category, _ = BookCategory.objects.get_or_create(
            category_name="Books",
            organization=organization,
            batch=batch,
            defaults={
                "category_description": "Books",
                "is_active": True,
                "created_by": created_by,
                "updated_by": created_by,
            },
        )

        subcategory, _ = BookSubCategory.objects.get_or_create(
            category=category,
            sub_category_name="Academic books",
            defaults={
                "sub_category_description": "Academic books",
                "is_active": True,
                "created_by": created_by,
                "updated_by": created_by,
            },
        )

        branches = []
        for name in ("Branch1", "Branch2", "Branch3"):
            branch, _ = LibraryBranch.objects.get_or_create(
                library_branch_name=name,
                organization=organization,
                batch=batch,
                defaults={"is_active": True},
            )
            if not branch.is_active:
                branch.is_active = True
                branch.save(update_fields=["is_active"])
            branches.append(branch)

        locations = []
        for name in ("Location1", "Location2", "Location3"):
            location, _ = BookLocation.objects.get_or_create(
                book_location=name,
                organization=organization,
                batch=batch,
                defaults={
                    "book_location_desc": name,
                    "is_active": True,
                    "created_by": created_by,
                    "updated_by": created_by,
                },
            )
            if not location.is_active:
                location.is_active = True
                location.save(update_fields=["is_active"])
            locations.append(location)

        return {
            "category": category,
            "subcategory": subcategory,
            "branch": branches[0],
            "location": locations[0],
            "organization": organization,
            "batch": batch,
        }

    def read_grouped_books(self, excel_path: str) -> tuple[OrderedDict[tuple[str, ...], GroupedBook], dict[str, int]]:
        workbook = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
        if not workbook.sheetnames:
            raise CommandError("Workbook contains no sheets.")

        worksheet = workbook[workbook.sheetnames[0]]
        header = next(worksheet.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW, values_only=True), None)
        if not header:
            raise CommandError("Workbook header row is missing.")

        grouped_books: OrderedDict[tuple[str, ...], GroupedBook] = OrderedDict()
        title_signature_count: dict[str, set[tuple[str, ...]]] = {}
        non_numeric_pages = 0
        non_numeric_volumes = 0

        for row in worksheet.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
            if not any(value not in (None, "") for value in row):
                continue

            title = clean_text(row[0])
            if not title:
                continue

            publisher = clean_text(row[3])
            author = clean_text(row[4])
            publish_year = clean_text(row[5])
            volume_raw = row[6]
            isbn = clean_text(row[7])
            edition = clean_text(row[8])
            pages_raw = row[9]

            volume = parse_int_or_none(volume_raw)
            pages = parse_int_or_none(pages_raw)

            if clean_text(volume_raw) and volume is None:
                non_numeric_volumes += 1
            if clean_text(pages_raw) and pages is None:
                non_numeric_pages += 1

            signature = (
                normalize_key(title),
                normalize_key(publisher),
                normalize_key(author),
                normalize_key(publish_year),
                normalize_key(volume_raw),
                normalize_key(isbn),
                normalize_key(edition),
                normalize_key(pages_raw),
            )

            title_signature_count.setdefault(normalize_key(title), set()).add(signature)

            if signature not in grouped_books:
                grouped_books[signature] = GroupedBook(
                    title=title,
                    publisher=publisher,
                    author=author,
                    publish_year=publish_year,
                    volume=volume,
                    isbn=isbn,
                    edition=edition,
                    pages=pages,
                    copy_count=0,
                )

            grouped_books[signature].copy_count += 1

        anomaly_summary = {
            "titles_with_multiple_signatures": sum(1 for signatures in title_signature_count.values() if len(signatures) > 1),
            "non_numeric_pages": non_numeric_pages,
            "non_numeric_volumes": non_numeric_volumes,
        }
        return grouped_books, anomaly_summary

    def import_books(
        self,
        grouped_books: OrderedDict[tuple[str, ...], GroupedBook],
        academic_year: AcademicYear,
        defaults: dict[str, Any],
        created_by: int,
    ) -> dict[str, int]:
        today = date.today()
        next_barcode = 1
        books_created = 0
        purchases_created = 0
        barcodes_created = 0

        for bill_no, book in enumerate(grouped_books.values(), start=1):
            with transaction.atomic():
                library_book = LibraryBook.objects.create(
                    book_code=f"BOOK{bill_no:05d}",
                    book_name=book.title,
                    book_category=defaults["category"],
                    book_sub_category=defaults["subcategory"],
                    library_branch=defaults["branch"],
                    book_status="ACTIVE",
                    no_of_copies=book.copy_count,
                    organization=defaults["organization"],
                    batch=defaults["batch"],
                    publisher=book.publisher or None,
                    author=book.author or None,
                    publish_year=book.publish_year or None,
                    volume=book.volume,
                    edition=book.edition or None,
                    pages=book.pages,
                    barcode_auto_generated=True,
                    ISBN=book.isbn or None,
                    academic_year=academic_year,
                    createdDate=today,
                    allow_issue="T",
                    type="book",
                    IssueNo="ISS123",
                    Period="Annual",
                    created_by=created_by,
                    updated_by=created_by,
                )
                books_created += 1

                LibraryPurchase.objects.create(
                    book=library_book,
                    purchase_date=today,
                    purchase_from="book stroe",
                    bill_no=str(bill_no),
                    bill_value=0,
                    bill_concession=0,
                    no_of_copies=book.copy_count,
                    created_by=created_by,
                    updated_by=created_by,
                )
                purchases_created += 1

                barcode_rows = []
                for _ in range(book.copy_count):
                    barcode_rows.append(
                        LibraryBooksBarcode(
                            book=library_book,
                            barcode=next_barcode,
                            book_barcode_status="ACTIVE",
                            remarks="",
                            barcode_auto_generated=True,
                            organization=defaults["organization"],
                            batch=defaults["batch"],
                            location_id=defaults["location"],
                            created_by=created_by,
                            updated_by=created_by,
                        )
                    )
                    next_barcode += 1
                LibraryBooksBarcode.objects.bulk_create(barcode_rows, batch_size=1000)
                barcodes_created += len(barcode_rows)

        return {
            "library_books": books_created,
            "library_purchases": purchases_created,
            "library_barcodes": barcodes_created,
        }
