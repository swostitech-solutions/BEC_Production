import os
import re
from pathlib import Path

import openpyxl

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "Swostitech_Acadix.settings")

import django

django.setup()

from django.db import connection, transaction

from Acadix.models import Address, City, Country, EmployeeMaster, State, StudentRegistration, UserLogin
from seed_staff_basic_address import WORKBOOK_PATH, clean_text, parse_address_blob


STUDENT_SOURCE_ROOT = Path(r"C:\Users\dhruv\Downloads\BEC_Student Data\BEC_Student Data")
STAFF_OLD_ID_START = 13
STAFF_OLD_ID_END = 53


def normalize_registration_no(value):
    if value is None:
        return None
    return re.sub(r"\s+", "", str(value).strip())


def normalize_phone(value):
    if value is None:
        return ""
    digits = re.sub(r"\D", "", str(value))
    return digits if len(digits) == 10 else ""


def load_staff_workbook():
    wb = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    rows = {}
    for r in range(2, ws.max_row + 1):
        values = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if not any(v not in (None, "") for v in values):
            continue
        row = {str(headers[i]): values[i] for i in range(len(headers)) if headers[i] is not None}
        employee_code = clean_text(row.get("Employee Code"))
        if employee_code:
            rows[employee_code] = row
    return rows


def load_student_source_addresses(target_registration_nos):
    found = {}
    for path in sorted(STUDENT_SOURCE_ROOT.glob("*.xlsx")):
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for ws in wb.worksheets:
            header_row = next(ws.iter_rows(values_only=True), None)
            if not header_row:
                continue
            headers = {str(v).strip(): i for i, v in enumerate(header_row) if v is not None}
            required = [
                "BPUT Registration No",
                "Present Address",
                "Present District",
                "Present State",
                "Present Country",
                "Present Pincode",
                "Student Phone",
            ]
            if any(h not in headers for h in required):
                continue
            for row in ws.iter_rows(min_row=2, values_only=True):
                reg = normalize_registration_no(row[headers["BPUT Registration No"]])
                if not reg or reg not in target_registration_nos or reg in found:
                    continue
                present_address = clean_text(row[headers["Present Address"]]) or ""
                present_city = clean_text(row[headers["Present District"]]) or ""
                present_state = clean_text(row[headers["Present State"]]) or ""
                present_country = clean_text(row[headers["Present Country"]]) or ""
                present_pincode = clean_text(row[headers["Present Pincode"]]) or ""
                phone = normalize_phone(row[headers["Student Phone"]])
                found[reg] = {
                    "present_address": present_address,
                    "present_city": present_city,
                    "present_state": present_state,
                    "present_country": present_country,
                    "present_pincode": present_pincode,
                    "present_phone_number": phone,
                    "permanent_address": present_address,
                    "permanent_city": present_city,
                    "permanent_state": present_state,
                    "permanent_country": present_country,
                    "permanent_pincode": present_pincode,
                    "permanent_phone_number": phone,
                }
                if len(found) == len(target_registration_nos):
                    return found
    return found


def set_employee_pk(old_id, new_id):
    with connection.cursor() as cursor:
        cursor.execute(
            'UPDATE "Acadix_employeemaster" SET id = %s WHERE id = %s',
            [new_id, old_id],
        )


def set_userlogin_reference(old_id, new_id):
    with connection.cursor() as cursor:
        cursor.execute(
            'UPDATE "Acadix_userlogin" SET reference_id = %s WHERE reference_id = %s AND user_type_id = 3',
            [new_id, old_id],
        )


def bump_employee_sequence():
    with connection.cursor() as cursor:
        cursor.execute(
            """SELECT setval(
                pg_get_serial_sequence('"Acadix_employeemaster"', 'id'),
                (SELECT COALESCE(MAX(id), 1) FROM "Acadix_employeemaster"),
                true
            )"""
        )


def main():
    staff_rows = load_staff_workbook()
    staff_qs = EmployeeMaster.objects.filter(id__gte=STAFF_OLD_ID_START, id__lte=STAFF_OLD_ID_END).order_by("id")
    old_staff = list(staff_qs.values("id", "employee_code"))
    if len(old_staff) != 41:
        raise ValueError(f"Expected 41 imported staff rows in id range {STAFF_OLD_ID_START}-{STAFF_OLD_ID_END}, found {len(old_staff)}")

    student_rows = list(
        StudentRegistration.objects.filter(id__gte=STAFF_OLD_ID_START, id__lte=STAFF_OLD_ID_END)
        .order_by("id")
        .values("id", "registration_no")
    )
    if len(student_rows) != 41:
        raise ValueError("Expected 41 student rows in overlapping id range.")

    target_regs = {row["registration_no"] for row in student_rows}
    source_addresses = load_student_source_addresses(target_regs)
    if len(source_addresses) != len(target_regs):
        missing = sorted(target_regs - set(source_addresses))
        raise ValueError(f"Could not restore all student source addresses. Missing: {missing}")

    new_id_start = (StudentRegistration.objects.order_by("-id").values_list("id", flat=True).first() or 0) + 1
    id_mapping = {row["id"]: new_id_start + index for index, row in enumerate(old_staff)}

    india = Country.objects.get(country_name__iexact="India", is_active=True)
    odisha = State.objects.get(state_name__iexact="Odisha", is_active=True)
    bhubaneswar = City.objects.get(city_name__iexact="Bhubaneswar", is_active=True)

    with transaction.atomic():
        for old_id in sorted(id_mapping.keys(), reverse=True):
            set_userlogin_reference(old_id, id_mapping[old_id])
            set_employee_pk(old_id, id_mapping[old_id])

        bump_employee_sequence()

        for student in student_rows:
            address_payload = source_addresses[student["registration_no"]]
            address = Address.objects.get(reference_id=student["id"])
            address.usertype = "STUDENT"
            address.present_address = address_payload["present_address"]
            address.present_pincode = address_payload["present_pincode"]
            address.present_city = address_payload["present_city"]
            address.present_state = address_payload["present_state"]
            address.present_country = address_payload["present_country"]
            address.present_phone_number = address_payload["present_phone_number"]
            address.permanent_address = address_payload["permanent_address"]
            address.permanent_pincode = address_payload["permanent_pincode"]
            address.permanent_city = address_payload["permanent_city"]
            address.permanent_state = address_payload["permanent_state"]
            address.permanent_country = address_payload["permanent_country"]
            address.permanent_phone_number = address_payload["permanent_phone_number"]
            address.updated_by = 1
            address.save()

        for old_id, new_id in id_mapping.items():
            employee = EmployeeMaster.objects.get(id=new_id)
            source_row = staff_rows[employee.employee_code]
            address_parts = parse_address_blob(source_row.get("Address"))
            mobile = normalize_phone(source_row.get("Mobile No")) or "1234567890"
            Address.objects.create(
                reference_id=new_id,
                organization=employee.organization,
                branch=employee.branch,
                usertype=employee.employee_type.employee_type_code,
                present_address=address_parts["address"],
                present_pincode=address_parts["pincode"],
                present_city=str(bhubaneswar.id),
                present_state=str(odisha.id),
                present_country=str(india.id),
                present_phone_number=mobile,
                permanent_address=address_parts["address"],
                permanent_pincode=address_parts["pincode"],
                permanent_city=str(bhubaneswar.id),
                permanent_state=str(odisha.id),
                permanent_country=str(india.id),
                permanent_phone_number=mobile,
                created_by=1,
                updated_by=1,
                is_active=True,
            )

    print(
        f"REPAIRED staff_rows={len(id_mapping)} new_id_start={new_id_start} new_id_end={max(id_mapping.values())}"
    )


if __name__ == "__main__":
    main()
