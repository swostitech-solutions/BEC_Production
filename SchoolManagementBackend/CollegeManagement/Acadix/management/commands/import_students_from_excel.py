import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Iterable, Optional

from django.contrib.auth.hashers import make_password
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from Acadix.models import (
    AcademicYear,
    Address,
    Batch,
    Blood,
    Course,
    Department,
    FeeFrequency,
    FeeStructureDetail,
    FeeStructureMaster,
    Gender,
    House,
    MotherTongue,
    Nationality,
    Parent,
    Religion,
    Section,
    Semester,
    StudentCourse,
    StudentFeeDetail,
    StudentRegistration,
    UserLogin,
    UserType,
)


EXPECTED_HEADERS = [
    "Session",
    "Course",
    "Department",
    "Academic Year",
    "Semester",
    "Section",
    "BPUT Registration No",
    "First Name",
    "Middle Name",
    "Last Name",
    "Date Of Birth",
    "Date Of Admission",
    "Gender",
    "Admission Type",
    "Student Phone",
    "Email",
    "Blood Group",
    "Religion",
    "Nationality",
    "Mother Tongue",
    "Present Address",
    "Present District",
    "Present State",
    "Present Country",
    "Present Pincode",
    "Father Name",
    "Father Contact Number",
    "Mother Name",
    "Mother Contact Number",
]


NULL_MARKERS = {"", "NULL", "NONE", "N/A", "NA", "-"}
BLANK_STREAK_BREAK = 50


@dataclass
class ImportStats:
    files: int = 0
    sheets: int = 0
    rows_seen: int = 0
    rows_skipped_blank: int = 0
    rows_skipped_existing: int = 0
    rows_imported: int = 0
    rows_failed: int = 0


class Command(BaseCommand):
    help = "Import student registrations from one Excel file or a folder of Excel files."

    def add_arguments(self, parser):
        parser.add_argument("path", type=str, help="Path to a .xlsx/.xls file or a folder containing Excel files.")
        parser.add_argument("--dry-run", action="store_true", help="Validate rows without writing to the DB.")
        parser.add_argument(
            "--created-by",
            type=int,
            default=1,
            help="User id to stamp into created_by/updated_by fields. Defaults to 1.",
        )
        parser.add_argument(
            "--limit",
            type=int,
            default=None,
            help="Optional maximum number of data rows to process across all files.",
        )

    def handle(self, *args, **options):
        root = Path(options["path"]).expanduser()
        if not root.exists():
            raise CommandError(f"Path not found: {root}")

        self.dry_run = options["dry_run"]
        self.created_by = options["created_by"]
        self.limit = options["limit"]
        self.stats = ImportStats()
        self.errors = []
        self.student_user_type = self._get_or_create_student_user_type()
        self.default_house = House.objects.filter(is_active=True).order_by("id").first()

        excel_files = self._collect_excel_files(root)
        if not excel_files:
            raise CommandError(f"No Excel files found under: {root}")

        self.stdout.write(
            self.style.NOTICE(
                f"{'Dry run' if self.dry_run else 'Import'} starting for {len(excel_files)} file(s)."
            )
        )

        for file_path in excel_files:
            if self.limit is not None and self.stats.rows_seen >= self.limit:
                break
            self._process_workbook(file_path)

        self._print_summary()
        if self.errors:
            self.stdout.write(self.style.WARNING("\nSample errors:"))
            for item in self.errors[:20]:
                self.stdout.write(
                    f"- {item['file']} | {item['sheet']} | row {item['row']}: {item['error']}"
                )

    def _collect_excel_files(self, root: Path) -> list[Path]:
        if root.is_file():
            return [root]
        return sorted(
            [
                path
                for path in root.rglob("*")
                if path.is_file() and path.suffix.lower() in {".xlsx", ".xls"}
            ]
        )

    def _get_or_create_student_user_type(self) -> UserType:
        user_type = UserType.objects.filter(user_type__iexact="student", is_active=True).first()
        if user_type:
            return user_type
        next_id = (UserType.objects.order_by("-id").values_list("id", flat=True).first() or 0) + 1
        return UserType.objects.create(
            id=next_id,
            user_type="student",
            description="student",
            created_by=self.created_by,
            updated_by=self.created_by,
            is_active=True,
        )

    def _process_workbook(self, file_path: Path) -> None:
        self.stats.files += 1
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        self.stdout.write(self.style.NOTICE(f"\nProcessing workbook: {file_path}"))
        workbook_hint = self._parse_workbook_hint(file_path)

        for worksheet in workbook.worksheets:
            if self.limit is not None and self.stats.rows_seen >= self.limit:
                break

            self.stats.sheets += 1
            header_map = self._extract_headers(worksheet)
            if header_map is None:
                self.stdout.write(f"  Sheet: {worksheet.title} (skipped empty)")
                continue
            self.stdout.write(f"  Sheet: {worksheet.title}")
            blank_streak = 0

            for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
                if row_index == 1:
                    continue
                if self.limit is not None and self.stats.rows_seen >= self.limit:
                    break

                row_dict = self._row_to_dict(row, header_map)
                if self._is_blank_row(row_dict):
                    self.stats.rows_skipped_blank += 1
                    blank_streak += 1
                    if blank_streak >= BLANK_STREAK_BREAK:
                        break
                    continue
                blank_streak = 0

                self.stats.rows_seen += 1
                try:
                    payload = self._prepare_row_payload(worksheet.title, row_dict, workbook_hint)
                    if not self.dry_run:
                        with transaction.atomic():
                            self._import_student(payload)
                    self.stats.rows_imported += 1
                except Exception as exc:  # noqa: BLE001
                    self.stats.rows_failed += 1
                    self.errors.append(
                        {
                            "file": str(file_path),
                            "sheet": worksheet.title,
                            "row": row_index,
                            "error": str(exc),
                        }
                    )

    def _extract_headers(self, worksheet) -> Optional[dict[str, int]]:
        header_row = next(worksheet.iter_rows(values_only=True), None)
        if not header_row:
            return None

        headers = {str(value).strip(): idx for idx, value in enumerate(header_row) if value is not None}
        missing = [header for header in EXPECTED_HEADERS if header not in headers]
        if missing:
            raise CommandError(
                f"Sheet '{worksheet.title}' is missing required columns: {', '.join(missing)}"
            )
        return headers

    def _row_to_dict(self, row: tuple, header_map: dict[str, int]) -> dict[str, object]:
        data = {}
        for header, idx in header_map.items():
            data[header] = row[idx] if idx < len(row) else None
        return data

    def _is_blank_row(self, row_dict: dict[str, object]) -> bool:
        return not any(self._normalize_scalar(value) is not None for value in row_dict.values())

    def _prepare_row_payload(self, sheet_name: str, row: dict[str, object], workbook_hint: dict[str, str | None]) -> dict[str, object]:
        batch_code = self._normalize_batch_code(row.get("Session")) or self._normalize_batch_code(sheet_name)
        course_code = self._normalize_course_code(row.get("Course") or workbook_hint.get("course"))
        department_name = self._normalize_department_name(row.get("Department") or workbook_hint.get("department"))
        raw_academic_year = row.get("Academic Year")
        raw_semester = row.get("Semester")
        academic_year_code = self._normalize_academic_year_code(raw_academic_year)
        semester_code = self._normalize_semester_code(raw_semester)
        academic_year_code = self._align_academic_year_with_semester(academic_year_code, semester_code)
        section_label = self._normalize_section_label(row.get("Section"))

        batch = Batch.objects.get(batch_code=batch_code, is_active=True)
        course = Course.objects.get(batch=batch, course_code=course_code, is_active=True)
        department = self._resolve_department(batch, course, department_name)
        academic_year = AcademicYear.objects.get(
            batch=batch,
            course=course,
            department=department,
            academic_year_code=academic_year_code,
            is_active=True,
        )
        semester = Semester.objects.get(
            batch=batch,
            course=course,
            department=department,
            academic_year=academic_year,
            semester_code=semester_code,
            is_active=True,
        )
        admission_type = self._normalize_admission_type(row.get("Admission Type"))
        semester_queryset = Semester.objects.filter(
            batch=batch,
            course=course,
            department=department,
            is_active=True,
        ).order_by("display_order", "id")
        batch_years = re.findall(r"(\d{4})", str(batch.batch_code or ""))
        is_three_year_batch = len(batch_years) >= 2 and int(batch_years[1]) - int(batch_years[0]) == 3
        if admission_type == "LATERAL" or is_three_year_batch:
            first_semester = semester_queryset.filter(display_order__gte=3).first() or semester_queryset.first()
        else:
            first_semester = semester_queryset.first()
        if not first_semester:
            raise ValueError(
                f"No semester found for {batch.batch_code} / {course.course_code} / {department.department_description}"
            )

        section = self._resolve_section(batch, course, department, academic_year, semester, section_label)
        fee_group = FeeStructureMaster.objects.get(
            batch=batch,
            course=course,
            department=department,
            is_active=True,
        )
        FeeFrequency.objects.get(
            batch=batch,
            course=course,
            department=department,
            is_active=True,
        )

        registration_no = self._stringify_registration_no(row.get("BPUT Registration No"))
        first_name, middle_name, last_name = self._resolve_name_parts(
            row.get("First Name"),
            row.get("Middle Name"),
            row.get("Last Name"),
        )
        student_phone = self._clean_phone(row.get("Student Phone"))
        email = self._normalize_email(row.get("Email"))
        dob = self._normalize_date(row.get("Date Of Birth"))
        doa = self._normalize_date(row.get("Date Of Admission"))

        gender = self._resolve_optional_master(Gender, row.get("Gender"), ("gender_code", "gender_name"), self._normalize_gender_label)
        blood = self._resolve_optional_master(Blood, row.get("Blood Group"), ("blood_code", "blood_name"))
        religion = self._resolve_optional_master(
            Religion,
            row.get("Religion"),
            ("religion_code", "religion_name"),
            self._normalize_title_case,
        )
        nationality = self._resolve_optional_master(
            Nationality,
            row.get("Nationality"),
            ("nationality_code", "nationality_name"),
            self._normalize_title_case,
        )
        mother_tongue = self._resolve_optional_master(
            MotherTongue,
            row.get("Mother Tongue"),
            ("mother_tongue_code", "mother_tongue_name"),
            self._normalize_title_case,
        )

        address_payload = self._build_address_payload(row, student_phone)

        return {
            "batch": batch,
            "course": course,
            "department": department,
            "academic_year": academic_year,
            "semester": semester,
            "section": section,
            "first_semester": first_semester,
            "fee_group": fee_group,
            "registration_no": registration_no,
            "first_name": first_name,
            "middle_name": middle_name,
            "last_name": last_name,
            "date_of_birth": dob,
            "date_of_admission": doa,
            "admission_type": admission_type,
            "student_phone": student_phone,
            "email": email,
            "gender": gender,
            "blood": blood,
            "religion": religion,
            "nationality": nationality,
            "mother_tongue": mother_tongue,
            "father_name": self._clean_person_name(row.get("Father Name")),
            "father_contact_number": self._clean_phone(row.get("Father Contact Number")),
            "mother_name": self._clean_person_name(row.get("Mother Name")),
            "mother_contact_number": self._clean_phone(row.get("Mother Contact Number")),
            "address": address_payload,
        }

    def _import_student(self, payload: dict[str, object]) -> None:
        batch = payload["batch"]
        course = payload["course"]
        department = payload["department"]
        academic_year = payload["academic_year"]
        semester = payload["semester"]
        section = payload["section"]
        first_semester = payload["first_semester"]
        fee_group = payload["fee_group"]

        admission_no = self._next_admission_no()
        user_name = self._generate_unique_username(payload["first_name"], admission_no)

        student = StudentRegistration.objects.create(
            organization=batch.organization,
            branch=batch.branch,
            batch=batch,
            admission_type=payload["admission_type"],
            course=course,
            department=department,
            academic_year=academic_year,
            semester=semester,
            section=section,
            date_of_admission=payload["date_of_admission"],
            date_of_birth=payload["date_of_birth"],
            college_admission_no=str(admission_no),
            religion=payload["religion"],
            gender=payload["gender"],
            nationality=payload["nationality"],
            house=None,
            contact_no=payload["student_phone"],
            blood=payload["blood"],
            enrollment_no=None,
            barcode="",
            admission_no=str(admission_no),
            registration_no=payload["registration_no"],
            category=None,
            mother_tongue=payload["mother_tongue"],
            status="ACTIVE",
            email=payload["email"],
            children_in_family=None,
            student_aadhaar_no=None,
            user_name=user_name,
            remarks=None,
            referred_by=None,
            father_name=payload["father_name"],
            mother_name=payload["mother_name"],
            father_contact_number=payload["father_contact_number"],
            mother_contact_number=payload["mother_contact_number"],
            created_by=self.created_by,
            updated_by=self.created_by,
            first_name=payload["first_name"],
            middle_name=payload["middle_name"],
            last_name=payload["last_name"],
        )

        student_course = StudentCourse.objects.create(
            organization=batch.organization,
            branch=batch.branch,
            academic_year=academic_year,
            student=student,
            batch=batch,
            course=course,
            department=department,
            semester=semester,
            section=section,
            fee_group=fee_group,
            fee_applied_from=first_semester,
            enrollment_no=None,
            house=self.default_house,
            transport_availed=False,
            route_id=None,
            choice_semester="",
            student_status="ACTIVE",
            created_by=self.created_by,
            updated_by=self.created_by,
        )

        self._create_student_fee_rows(student, student_course, fee_group)
        self._create_address(student, payload["address"])
        Parent.objects.create(parent_id=self._next_parent_id(), student=student, is_active=True)

        user_login = UserLogin.objects.create(
            user_name=user_name,
            plain_password=payload["first_name"],
            reference_id=student.id,
            user_type=self.student_user_type,
            organization=batch.organization,
            branch=batch.branch,
            is_active=True,
            is_staff=False,
            password=make_password(payload["first_name"]),
        )
        user_login.save(update_fields=["password"])

    def _create_student_fee_rows(
        self,
        student: StudentRegistration,
        student_course: StudentCourse,
        fee_group: FeeStructureMaster,
    ) -> None:
        detail_rows = FeeStructureDetail.objects.filter(fee_structure_master=fee_group, is_active=True)
        fee_applied_from = student_course.fee_applied_from
        if not fee_applied_from:
            raise ValueError(f"Student course {student_course.id} is missing fee_applied_from")

        start_order = fee_applied_from.display_order or fee_applied_from.id or 0

        for detail in detail_rows:
            periods = detail.element_frequency.frequency_period
            semester_ids = [
                detail.semester_1,
                detail.semester_2,
                detail.semester_3,
                detail.semester_4,
                detail.semester_5,
                detail.semester_6,
                detail.semester_7,
                detail.semester_8,
            ]
            selected_semester_ids = semester_ids[:periods]
            if any(not semester_id for semester_id in selected_semester_ids):
                raise ValueError(
                    f"Fee structure detail {detail.id} is missing one of the semester slots for period {periods}"
                )

            for semester_id in selected_semester_ids:
                semester = Semester.objects.get(id=semester_id, is_active=True)
                semester_order = semester.display_order or semester.id or 0
                if semester_order < start_order:
                    continue

                StudentFeeDetail.objects.create(
                    student=student,
                    student_course=student_course,
                    fee_group=fee_group,
                    fee_structure_details=detail,
                    element_name=detail.element_type.element_name,
                    fee_applied_from=fee_applied_from,
                    semester=semester,
                    paid="N",
                    organization=fee_group.organization,
                    branch=fee_group.branch,
                    academic_year=student.academic_year,
                    department=fee_group.department,
                    multiplying_factor=1,
                    element_amount=detail.amount,
                    total_element_period_amount=detail.amount,
                    paid_amount=0,
                    created_by=self.created_by,
                    updated_by=self.created_by,
                )

    def _create_address(self, student: StudentRegistration, address_payload: dict[str, object]) -> None:
        if Address.objects.filter(reference_id=student.id, usertype="STUDENT", is_active=True).exists():
            return
        Address.objects.create(
            reference_id=student.id,
            organization=student.organization,
            branch=student.branch,
            usertype="STUDENT",
            present_address=address_payload["present_address"],
            present_pincode=address_payload["present_pincode"],
            present_city=address_payload["present_city"],
            present_state=address_payload["present_state"],
            present_country=address_payload["present_country"],
            present_phone_number=address_payload["present_phone_number"],
            permanent_address=address_payload["permanent_address"],
            permanent_pincode=address_payload["permanent_pincode"],
            permanent_city=address_payload["permanent_city"],
            permanent_state=address_payload["permanent_state"],
            permanent_country=address_payload["permanent_country"],
            permanent_phone_number=address_payload["permanent_phone_number"],
            created_by=self.created_by,
            updated_by=self.created_by,
        )

    def _build_address_payload(self, row: dict[str, object], phone: Optional[str]) -> dict[str, object]:
        present_address = self._normalize_scalar(row.get("Present Address")) or ""
        present_city = self._normalize_title_case(row.get("Present District")) or ""
        present_state = self._normalize_title_case(row.get("Present State")) or ""
        present_country = self._normalize_title_case(row.get("Present Country")) or ""
        present_pincode = self._clean_pincode(row.get("Present Pincode")) or ""
        return {
            "present_address": present_address,
            "present_pincode": present_pincode,
            "present_city": present_city,
            "present_state": present_state,
            "present_country": present_country,
            "present_phone_number": phone or "",
            "permanent_address": present_address,
            "permanent_pincode": present_pincode,
            "permanent_city": present_city,
            "permanent_state": present_state,
            "permanent_country": present_country,
            "permanent_phone_number": phone or "",
        }

    def _next_admission_no(self) -> str:
        highest = 1000
        for value in StudentRegistration.objects.values_list("admission_no", flat=True):
            if value is None:
                continue
            digits = re.sub(r"\D", "", str(value))
            if digits:
                highest = max(highest, int(digits))
        return str(highest + 1)

    def _next_parent_id(self) -> int:
        last = Parent.objects.order_by("-parent_id").first()
        return (last.parent_id + 1) if last else 101

    def _parse_workbook_hint(self, file_path: Path) -> dict[str, str | None]:
        stem = file_path.stem.upper()
        compact = re.sub(r"[^A-Z0-9]", "", stem)

        course_hint = None
        if "DIPLOMA" in compact or "DIPLOMA" in stem:
            course_hint = "DIPLOMA"
        elif "MBA" in compact:
            course_hint = "MBA"
        else:
            course_hint = "BTECH"

        department_hint = None
        file_department_map = {
            "AERONAUTICALBETECH": "Aeronautical Engineering",
            "AGRICULTUREBETECH": "Agriculture Engineering",
            "BTECHAME": "Aircraft Maintenance Engineering",
            "BTECHCIVILENVIRONMENT": "Civil & Environmental Engineering",
            "BTECHCIVIL": "Civil Engineering",
            "BTECHCOMPUTERSCIENCEENGINEERINGDATASCIENCE": "Computer Science & Engineering (Data Science)",
            "BTECHELECTRIALEENGINEERING": "Electrical Engineering",
            "BTECHELECTRIALEENGINEERING": "Electrical Engineering",
            "BTECHMECHANICALANDMECHATRONICS": "Mechanical & Mechatronics Engineering (Additive Manufacturing)",
            "COMPUTERSCIENCEENGINEERINGBETECH": "Computer Science & Engineering",
            "DIPLOMAAERONAUTICAL": "Aeronautical Engineering",
            "DIPLOMAAME": "Aircraft Maintenance Engineering",
            "DIPLOMACIVIL": "Civil Engineering",
            "ELECTRICALANDCOMPUTERBETECH": "Electrical & Computer Engineering",
            "ELECTRICALDIPLOMA": "Electrical Engineering",
            "MBABEC": "General",
            "MECHANICALBETECH": "Mechanical Engineering",
            "MECHANICALDIPLOMA": "Mechanical Engineering",
        }
        department_hint = file_department_map.get(compact)
        return {"course": course_hint, "department": department_hint}

    def _generate_unique_username(self, first_name: str, admission_no: str) -> str:
        base = re.sub(r"[^A-Za-z0-9]", "", first_name or "student").lower() or "student"
        candidate = f"{base}{admission_no}"
        suffix = 1
        while UserLogin.objects.filter(user_name=candidate).exists():
            suffix += 1
            candidate = f"{base}{admission_no}{suffix}"
        return candidate

    def _resolve_department(self, batch: Batch, course: Course, raw_name: str) -> Department:
        departments = list(
            Department.objects.filter(batch=batch, course=course, is_active=True).order_by("id")
        )
        alias_map = {}
        for department in departments:
            for candidate in {
                department.department_description,
                department.department_code,
                self._department_abbreviation(department.department_description),
            }:
                if candidate:
                    alias_map[self._slugify(candidate)] = department
        key = self._slugify(raw_name)
        if key in alias_map:
            return alias_map[key]
        raise Department.DoesNotExist(
            f"Department '{raw_name}' not found for {batch.batch_code} / {course.course_code}"
        )

    def _resolve_section(
        self,
        batch: Batch,
        course: Course,
        department: Department,
        academic_year: AcademicYear,
        semester: Semester,
        section_label: str,
    ) -> Section:
        sections = Section.objects.filter(
            batch=batch,
            course=course,
            department=department,
            academic_year=academic_year,
            semester=semester,
            is_active=True,
        )
        for section in sections:
            if self._slugify(section.section_code) == self._slugify(section_label):
                return section
            if self._slugify(section.section_name) == self._slugify(section_label):
                return section
        raise Section.DoesNotExist(
            f"Section '{section_label}' not found for {batch.batch_code} / {course.course_code} / "
            f"{department.department_description} / {semester.semester_code}"
        )

    def _resolve_optional_master(
        self,
        model,
        raw_value: object,
        fields: Iterable[str],
        normalizer=None,
    ):
        value = self._normalize_scalar(raw_value)
        if value is None:
            return None
        normalized = normalizer(value) if normalizer else value
        queryset = model.objects.filter(is_active=True)
        for field in fields:
            instance = queryset.filter(**{f"{field}__iexact": normalized}).first()
            if instance:
                return instance
        raise model.DoesNotExist(f"{model.__name__} value '{value}' not found.")

    def _normalize_batch_code(self, value: object) -> str:
        normalized = self._normalize_scalar(value)
        if not normalized:
            raise ValueError("Session/Batch is required.")
        digits = re.sub(r"[^0-9]", "", normalized)
        if len(digits) == 8:
            normalized = f"{digits[:4]}-{digits[4:]}"
        return normalized

    def _normalize_course_code(self, value: object) -> str:
        normalized = self._slugify(self._normalize_scalar(value))
        course_map = {
            "BTECH": "BTECH",
            "BTECHNOLOGY": "BTECH",
            "DIPLOMA": "DIPLOMA",
            "DIP": "DIPLOMA",
            "MBA": "MBA",
        }
        if normalized in course_map:
            return course_map[normalized]
        raise ValueError(f"Unsupported course value: {value}")

    def _normalize_department_name(self, value: object) -> str:
        normalized = self._normalize_scalar(value)
        if not normalized:
            raise ValueError("Department is required.")
        dept_aliases = {
            "MECHANICAL": "Mechanical Engineering",
            "MECHANICALENGINEERING": "Mechanical Engineering",
            "CIVIL": "Civil Engineering",
            "ELECTRICAL": "Electrical Engineering",
            "ELECTRIALENGINEERING": "Electrical Engineering",
            "ELECTRIAL": "Electrical Engineering",
            "AERONAUTICAL": "Aeronautical Engineering",
            "AIRCRAFTMAINTENANCE": "Aircraft Maintenance Engineering",
            "AME": "Aircraft Maintenance Engineering",
            "GENERAL": "General",
            "COMPUTERSCIENCEENGINEERING": "Computer Science & Engineering",
            "COMPUTERSCIENCEENGINEERINGDATASCIENCE": "Computer Science & Engineering (Data Science)",
            "CIVILENVIRONMENT": "Civil & Environmental Engineering",
            "ELECTRICALANDCOMPUTER": "Electrical & Computer Engineering",
            "MECHANICALANDMECHATRONICS": "Mechanical & Mechatronics Engineering (Additive Manufacturing)",
            "AGRICULTURE": "Agriculture Engineering",
        }
        return dept_aliases.get(self._slugify(normalized), normalized)

    def _normalize_academic_year_code(self, value: object) -> str:
        normalized = self._normalize_scalar(value)
        if not normalized:
            raise ValueError("Academic Year is required.")
        match = re.search(r"(\d+)", normalized)
        if not match:
            raise ValueError(f"Could not understand academic year value: {value}")
        year_no = int(match.group(1))
        ordinal = {1: "1st", 2: "2nd", 3: "3rd", 4: "4th", 5: "5th"}.get(year_no, f"{year_no}th")
        return f"{ordinal} Year"

    def _normalize_semester_code(self, value: object) -> str:
        normalized = self._normalize_scalar(value)
        if not normalized:
            raise ValueError("Semester is required.")
        normalized = normalized.replace("Smester", "Semester").replace("smester", "semester")
        match = re.search(r"(\d+)", normalized)
        if not match:
            raise ValueError(f"Could not understand semester value: {value}")
        sem_no = int(match.group(1))
        ordinal = {1: "1st", 2: "2nd", 3: "3rd"}.get(sem_no, f"{sem_no}th")
        return f"{ordinal} Semester"

    def _align_academic_year_with_semester(self, academic_year_code: str, semester_code: str) -> str:
        year_match = re.search(r"(\d+)", academic_year_code)
        sem_match = re.search(r"(\d+)", semester_code)
        if not year_match or not sem_match:
            return academic_year_code
        current_year_no = int(year_match.group(1))
        semester_no = int(sem_match.group(1))
        derived_year_no = ((semester_no - 1) // 2) + 1
        if current_year_no == derived_year_no:
            return academic_year_code
        ordinal = {1: "1st", 2: "2nd", 3: "3rd", 4: "4th", 5: "5th"}.get(derived_year_no, f"{derived_year_no}th")
        return f"{ordinal} Year"

    def _normalize_section_label(self, value: object) -> str:
        normalized = self._normalize_scalar(value)
        if not normalized:
            return "A"
        if self._slugify(normalized) == "SECTIONA":
            return "A"
        return normalized

    def _normalize_gender_label(self, value: str) -> str:
        key = self._slugify(value)
        if key == "M":
            return "M"
        if key == "MALE":
            return "Male"
        if key == "F":
            return "F"
        if key == "FEMALE":
            return "Female"
        return value

    def _normalize_admission_type(self, value: object) -> str:
        normalized = (self._normalize_scalar(value) or "REGULAR").upper()
        if normalized == "LATERAL":
            return "LATERAL"
        return "REGULAR"

    def _normalize_date(self, value: object) -> Optional[date]:
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        normalized = self._normalize_scalar(value)
        if normalized is None:
            return None
        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(normalized, fmt).date()
            except ValueError:
                continue
        raise ValueError(f"Invalid date value: {value}")

    def _normalize_email(self, value: object) -> Optional[str]:
        normalized = self._normalize_scalar(value)
        if not normalized:
            return None
        return normalized.lower()

    def _stringify_registration_no(self, value: object) -> str:
        normalized = self._normalize_scalar(value)
        if not normalized:
            return ""
        return str(normalized)

    def _clean_person_name(self, value: object, required: bool = False) -> Optional[str]:
        normalized = self._normalize_scalar(value)
        if normalized is None:
            if required:
                raise ValueError("Required person name is missing.")
            return None
        cleaned = re.sub(r"\s+", " ", normalized).strip()
        if required and not cleaned:
            raise ValueError("Required person name is blank.")
        return cleaned or None

    def _resolve_name_parts(
        self,
        first_value: object,
        middle_value: object,
        last_value: object,
    ) -> tuple[str, Optional[str], Optional[str]]:
        first_name = self._clean_person_name(first_value)
        middle_name = self._clean_person_name(middle_value)
        last_name = self._clean_person_name(last_value)

        if first_name and not middle_name and not last_name:
            tokens = first_name.split()
            if len(tokens) == 1:
                return tokens[0], None, None
            if len(tokens) == 2:
                return tokens[0], None, tokens[1]
            return tokens[0], " ".join(tokens[1:-1]), tokens[-1]

        if first_name and middle_name and not last_name:
            middle_tokens = middle_name.split()
            if len(middle_tokens) == 1:
                return first_name, None, middle_tokens[0]
            return first_name, " ".join(middle_tokens[:-1]), middle_tokens[-1]

        if not first_name and last_name:
            tokens = last_name.split()
            if len(tokens) == 1:
                raise ValueError("Student first name is missing.")
            if len(tokens) == 2:
                return tokens[0], None, tokens[1]
            return tokens[0], " ".join(tokens[1:-1]), tokens[-1]

        if not first_name:
            raise ValueError("Student first name is missing.")

        return first_name, middle_name, last_name

    def _clean_phone(self, value: object) -> Optional[str]:
        normalized = self._normalize_scalar(value)
        if normalized is None:
            return None
        digits = re.sub(r"\D", "", normalized)
        if len(digits) == 10:
            return digits
        return None

    def _clean_pincode(self, value: object) -> Optional[str]:
        normalized = self._normalize_scalar(value)
        if normalized is None:
            return None
        digits = re.sub(r"\D", "", normalized)
        if len(digits) == 6:
            return digits
        return None

    def _normalize_title_case(self, value: object) -> Optional[str]:
        normalized = self._normalize_scalar(value)
        if normalized is None:
            return None
        return re.sub(r"\s+", " ", normalized).strip().title()

    def _normalize_scalar(self, value: object) -> Optional[str]:
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.date().isoformat()
        if isinstance(value, date):
            return value.isoformat()
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        text = str(value).strip()
        if text.upper() in NULL_MARKERS:
            return None
        return text

    def _department_abbreviation(self, value: str) -> str:
        tokens = re.findall(r"[A-Za-z0-9]+", value.upper())
        return "".join(tokens)

    def _slugify(self, value: object) -> str:
        normalized = self._normalize_scalar(value) or ""
        return re.sub(r"[^A-Z0-9]", "", normalized.upper())

    def _print_summary(self) -> None:
        style = self.style.SUCCESS if self.stats.rows_failed == 0 else self.style.WARNING
        self.stdout.write(style("\nImport summary"))
        self.stdout.write(f"- files: {self.stats.files}")
        self.stdout.write(f"- sheets: {self.stats.sheets}")
        self.stdout.write(f"- rows_seen: {self.stats.rows_seen}")
        self.stdout.write(f"- rows_skipped_blank: {self.stats.rows_skipped_blank}")
        self.stdout.write(f"- rows_skipped_existing: {self.stats.rows_skipped_existing}")
        self.stdout.write(f"- rows_imported: {self.stats.rows_imported}")
        self.stdout.write(f"- rows_failed: {self.stats.rows_failed}")
